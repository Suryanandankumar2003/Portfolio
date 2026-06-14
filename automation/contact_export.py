#!/usr/bin/env python3
"""
contact_export.py — Export contacts.csv to Excel + summary report
Usage: python automation/contact_export.py
"""
import csv, json, sys, logging
from pathlib import Path
from datetime import datetime

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

CSV_FILE  = Path(__file__).parent.parent / "backend" / "contacts.csv"
OUT_DIR   = Path(__file__).parent.parent / "backend" / "data"

def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.warning("openpyxl not installed. Run: pip install openpyxl"); return

    if not CSV_FILE.exists():
        log.warning("No contacts.csv found."); return

    with open(CSV_FILE, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook(); ws = wb.active; ws.title = "Contacts"
    hdr_fill = PatternFill("solid", start_color="050A10")
    hdr_font = Font(name="Arial", bold=True, color="00E5FF", size=11)
    headers  = ["timestamp","name","email","subject","message"]

    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h.upper()); c.font=hdr_font; c.fill=hdr_fill
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        fill = PatternFill("solid", start_color="0A1628" if i%2==0 else "050A10")
        for j, h in enumerate(headers, 1):
            c = ws.cell(i, j, row.get(h,""))
            c.font = Font(name="Arial", size=9, color="E2E8F0"); c.fill=fill

    out = OUT_DIR / f"contacts_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(out); log.info(f"✓ Exported {len(rows)} contacts → {out}")

def print_summary():
    if not CSV_FILE.exists():
        log.info("No contacts yet."); return
    with open(CSV_FILE, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    log.info(f"\n{'='*50}\n  CONTACT SUMMARY — {datetime.now().strftime('%Y-%m-%d')}\n{'='*50}")
    log.info(f"  Total contacts: {len(rows)}")
    roles = {}
    for r in rows:
        roles[r.get("subject","Other")] = roles.get(r.get("subject","Other"),0)+1
    log.info(f"  Latest: {rows[-1].get('name','?')} <{rows[-1].get('email','?')}>" if rows else "  No contacts")

if __name__ == "__main__":
    print_summary(); export_excel()
