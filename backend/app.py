"""
app.py — FastAPI Backend v3
============================
POST /contact  →  1) contacts.csv  2) contacts.xlsx  3) contacts.db (SQLite)
POST /chat     →  RAG chatbot
GET  /health   →  status
"""
import os, csv, sqlite3, logging, logging.handlers, time
from datetime import datetime
from pathlib import Path
from collections import defaultdict

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel, validator
from dotenv import load_dotenv

load_dotenv()

# ── Logging ──────────────────────────────────────────────
LOG_DIR = Path("./logs"); LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.handlers.RotatingFileHandler(
            LOG_DIR / "portfolio.log", maxBytes=5_000_000, backupCount=3),
    ],
)
logger = logging.getLogger(__name__)

# ── Paths ─────────────────────────────────────────────────
BASE         = Path(__file__).parent
CONTACTS_CSV  = BASE / "contacts.csv"
CONTACTS_XLSX = BASE / "contacts.xlsx"
CONTACTS_DB   = BASE / "contacts.db"

# ── FastAPI ───────────────────────────────────────────────
app = FastAPI(
    title="SK Portfolio API",
    description="Suryanandan Kumar — DevOps · Cloud · AI Portfolio Backend",
    version="3.0.0", docs_url="/docs", redoc_url="/redoc",
)
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_credentials=True,
    allow_methods=["GET","POST","OPTIONS"], allow_headers=["*"],
)

# ── Rate limiter ──────────────────────────────────────────
_hits: dict = defaultdict(list)
def _rate_limit(ip: str, limit: int, window: int) -> bool:
    now = time.time()
    _hits[ip] = [t for t in _hits[ip] if now - t < window]
    if len(_hits[ip]) >= limit: return True
    _hits[ip].append(now); return False

# ── XSS sanitize ─────────────────────────────────────────
def _xss(v: str) -> str:
    for ch in ["<", ">", '"', "'"]: v = v.replace(ch, "")
    return v.strip()

# ── SQLite init ───────────────────────────────────────────
def _init_db():
    con = sqlite3.connect(CONTACTS_DB)
    con.execute("""
        CREATE TABLE IF NOT EXISTS contacts (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            name      TEXT NOT NULL,
            email     TEXT NOT NULL,
            subject   TEXT NOT NULL,
            message   TEXT NOT NULL
        )
    """)
    con.commit(); con.close()

# ── Excel writer ──────────────────────────────────────────
def _write_excel(rows):
    """Rewrite contacts.xlsx from all rows in DB."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook(); ws = wb.active; ws.title = "Contacts"
        thin = Border(
            left=Side(style="thin", color="1E3A5F"),
            right=Side(style="thin", color="1E3A5F"),
            top=Side(style="thin", color="1E3A5F"),
            bottom=Side(style="thin", color="1E3A5F"),
        )

        # Title row
        ws.merge_cells("A1:F1")
        ws["A1"] = "SURYANANDAN KUMAR — CONTACT SUBMISSIONS"
        ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="00E5FF")
        ws["A1"].fill      = PatternFill("solid", start_color="050A10")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # Headers
        hdrs = ["#", "Timestamp", "Name", "Email", "Subject", "Message"]
        hfill = PatternFill("solid", start_color="0A1628")
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(2, j, h)
            c.font      = Font(name="Arial", bold=True, size=10, color="00E5FF")
            c.fill      = hfill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin
        ws.row_dimensions[2].height = 24

        # Data rows
        for i, row in enumerate(rows, 3):
            fill = PatternFill("solid", start_color="0A1628" if i % 2 == 0 else "050A10")
            for j, val in enumerate(row, 1):
                c = ws.cell(i, j, val)
                c.font      = Font(name="Arial", size=9, color="E2E8F0")
                c.fill      = fill
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border    = thin
            ws.row_dimensions[i].height = 36

        # Column widths
        for col, w in [(1,5),(2,20),(3,22),(4,32),(5,28),(6,55)]:
            ws.column_dimensions[chr(64+col)].width = w

        # Summary row
        sr = len(rows) + 3
        ws.cell(sr, 1, f"Total: {len(rows)}").font = Font(name="Arial", bold=True, color="22C55E", size=9)

        wb.save(CONTACTS_XLSX)
        logger.info(f"Excel updated → {CONTACTS_XLSX} ({len(rows)} rows)")
    except Exception as e:
        logger.error(f"Excel write error: {e}")

# ── Pydantic models ───────────────────────────────────────
class ChatRequest(BaseModel):
    question:   str
    session_id: str = ""
    @validator("question")
    def clean(cls, v):
        v = _xss(v)
        if not v:       raise ValueError("question required")
        if len(v)>1000: raise ValueError("too long")
        return v

class ChatResponse(BaseModel):
    answer: str; session_id: str = ""; status: str = "ok"

class ContactRequest(BaseModel):
    name: str; email: str; subject: str; message: str
    @validator("name","subject","message")
    def clean_field(cls, v):
        v = _xss(v)
        if not v: raise ValueError("required")
        return v[:500]
    @validator("email")
    def clean_email(cls, v):
        v = _xss(v).lower()
        if "@" not in v or "." not in v: raise ValueError("invalid email")
        return v[:200]

class ContactResponse(BaseModel):
    message: str; status: str = "success"
    saved_to: list = []

# ── Startup ───────────────────────────────────────────────
@app.on_event("startup")
async def startup():
    logger.info("🚀 SK Portfolio API v3 starting...")
    _init_db()
    logger.info("✓ SQLite DB ready")
    try:
        from chatbot import get_manager; get_manager()
        logger.info("✓ Chatbot ready")
    except Exception as e:
        logger.warning(f"Chatbot init: {e}")

# ── Routes ────────────────────────────────────────────────
@app.get("/health")
def health():
    con = sqlite3.connect(CONTACTS_DB)
    count = con.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
    con.close()
    return {
        "status": "ok", "version": "3.0.0",
        "owner": "Suryanandan Kumar",
        "total_contacts": count,
        "timestamp": datetime.utcnow().isoformat(),
    }

@app.post("/chat", response_model=ChatResponse)
async def chat(req: ChatRequest, request: Request):
    ip = request.client.host if request.client else "unknown"
    if _rate_limit(ip, 30, 60):
        raise HTTPException(429, "Too many requests.")
    logger.info(f"[CHAT] {req.question[:80]}")
    try:
        from chatbot import get_manager
        r = get_manager().chat(req.question, req.session_id or None)
        return ChatResponse(answer=r["answer"], session_id=r.get("session_id",""))
    except Exception as e:
        logger.error(f"Chat error: {e}")
        return ChatResponse(
            answer="I'm having a moment. Email suryanandankumar2003@gmail.com",
            status="error")

@app.post("/contact", response_model=ContactResponse)
async def contact(req: ContactRequest, request: Request):
    ip = request.client.host if request.client else "unknown"
    if _rate_limit(ip, 5, 300):
        raise HTTPException(429, "Too many submissions. Wait 5 minutes.")

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    saved = []

    # ── 1. CSV ────────────────────────────────────────────
    try:
        exists = CONTACTS_CSV.exists() and CONTACTS_CSV.stat().st_size > 0
        with open(CONTACTS_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if not exists:
                w.writerow(["timestamp","name","email","subject","message"])
            w.writerow([ts, req.name, req.email, req.subject, req.message])
        saved.append("CSV")
        logger.info(f"[CSV] Saved: {req.name} <{req.email}>")
    except Exception as e:
        logger.error(f"CSV error: {e}")

    # ── 2. SQLite DB ──────────────────────────────────────
    try:
        con = sqlite3.connect(CONTACTS_DB)
        con.execute(
            "INSERT INTO contacts (timestamp,name,email,subject,message) VALUES (?,?,?,?,?)",
            (ts, req.name, req.email, req.subject, req.message)
        )
        con.commit()
        all_rows = con.execute(
            "SELECT id,timestamp,name,email,subject,message FROM contacts ORDER BY id"
        ).fetchall()
        con.close()
        saved.append("SQLite DB")
        logger.info(f"[DB] Saved contact #{len(all_rows)}")
    except Exception as e:
        logger.error(f"DB error: {e}")
        all_rows = []

    # ── 3. Excel ──────────────────────────────────────────
    try:
        _write_excel(all_rows)
        saved.append("Excel (.xlsx)")
    except Exception as e:
        logger.error(f"Excel error: {e}")

    logger.info(f"[CONTACT] ✓ Saved to: {', '.join(saved)}")
    return ContactResponse(
        message=f"Thanks {req.name}! Message received. Suryanandan will reply soon.",
        saved_to=saved,
    )

# ── Error handlers ────────────────────────────────────────
@app.exception_handler(404)
async def nf(req, exc): return JSONResponse(404, {"error":"Not found","docs":"/docs"})
@app.exception_handler(500)
async def se(req, exc): return JSONResponse(500, {"error":"Internal server error"})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app",
                host=os.getenv("HOST","0.0.0.0"),
                port=int(os.getenv("PORT",8000)),
                reload=True)
